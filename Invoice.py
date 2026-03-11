import os
import logging
import asyncio
from datetime import datetime
from aiohttp import web
import aiohttp
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, 
    CommandHandler, 
    MessageHandler, 
    filters, 
    ContextTypes, 
    ConversationHandler,
    CallbackQueryHandler
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
import uuid
import tempfile
import sys
import requests
from bs4 import BeautifulSoup
import re
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from googlesearch import search
import csv
from openpyxl import Workbook

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Conversation states
CLIENT_NAME, SERVICE, PRICE, QUANTITY, ADD_MORE, PAYMENT_DATE, PAYMENT_METHOD = range(7)

# New states for Lead Finder
MAIN_MENU, LEAD_FINDER_MENU, NICHE_LOCATION, UPLOAD_FILE, SENDER_EMAIL, EMAIL_PASSWORD, EMAIL_SUBJECT, EMAIL_BODY = range(7, 15)

# Company branding
COMPANY_NAME = "Tachaelhub"
COLORS = {
    'white': colors.white,
    'deep_blue': colors.HexColor('#1a237e'),
    'gold': colors.HexColor('#ffd700')
}

class InvoiceData:
    def __init__(self):
        self.client_name = ""
        self.items = []
        self.invoice_number = ""
        self.payment_date = ""
        self.payment_method = ""
        
    def reset(self):
        self.client_name = ""
        self.items = []
        self.invoice_number = ""
        self.payment_date = ""
        self.payment_method = ""
        
    def add_item(self, service, price, quantity):
        self.items.append({
            'service': service,
            'price': float(price),
            'quantity': int(quantity)
        })
        
    def calculate_total(self):
        return sum(item['price'] * item['quantity'] for item in self.items)
    
    def generate_invoice_number(self):
        date_str = datetime.now().strftime("%Y%m")
        unique_id = str(uuid.uuid4())[:8].upper()
        self.invoice_number = f"INV-{date_str}-{unique_id}"
        return self.invoice_number

# Store user data (in production, use a database)
user_data = {}

# Global application instance for webhook handler
application = None

# HTTP Server handlers
async def handle_health(request):
    return web.Response(text="healthy")

async def handle_root(request):
    return web.Response(text="Tachaelhub Invoice Bot is running!")

async def handle_webhook(request):
    """Handle incoming Telegram webhook updates"""
    global application
    try:
        # Parse the incoming update
        update_data = await request.json()
        update = Update.de_json(update_data, application.bot)
        
        # Process the update
        await application.process_update(update)
        
        return web.Response(status=200)
    except Exception as e:
        logger.error(f"Error processing webhook: {e}")
        return web.Response(status=500)

async def start_http_server():
    app = web.Application()
    app.router.add_get('/', handle_root)
    app.router.add_get('/health', handle_health)
    app.router.add_post('/webhook', handle_webhook)
    
    PORT = int(os.environ.get('PORT', 10000))
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, '0.0.0.0', PORT)
    await site.start()
    logger.info(f"HTTP server started on port {PORT}")

# Telegram handlers (keep all your existing handlers exactly as they are)
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start command handler - show main menu"""
    user_id = update.effective_user.id
    user_data[user_id] = InvoiceData()
    
    welcome_message = (
        f"🏢 Welcome to {COMPANY_NAME} Bot!\n\n"
        "Choose an option:"
    )
    
    keyboard = [
        [InlineKeyboardButton("📄 Generate Invoice", callback_data="generate_invoice")],
        [InlineKeyboardButton("🔍 Lead Finder", callback_data="lead_finder")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(welcome_message, reply_markup=reply_markup)
    return MAIN_MENU

async def main_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle main menu selections"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "generate_invoice":
        await query.edit_message_text(
            "🏢 *Invoice Generator*\n\n"
            "I'll help you create professional after-payment invoices.\n"
            "Let's start by entering the client information.\n\n"
            "Please enter the *client name*:",
            parse_mode='Markdown'
        )
        return CLIENT_NAME
    elif query.data == "lead_finder":
        keyboard = [
            [InlineKeyboardButton("🔍 Find Business Leads", callback_data="find_leads")],
            [InlineKeyboardButton("📧 Send Outreach Emails", callback_data="send_emails")],
            [InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_main")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            "🔍 *Lead Finder*\n\n"
            "Choose an option:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return LEAD_FINDER_MENU

async def lead_finder_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle lead finder menu selections"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "find_leads":
        await query.edit_message_text(
            "🔍 *Find Business Leads*\n\n"
            "Enter the niche and location in the format:\n"
            "*Niche Location*\n\n"
            "Examples:\n"
            "Schools Lagos\n"
            "Churches Abuja\n"
            "Restaurants Lekki",
            parse_mode='Markdown'
        )
        return NICHE_LOCATION
    elif query.data == "send_emails":
        await query.edit_message_text(
            "📧 *Send Outreach Emails*\n\n"
            "Please upload a CSV or XLSX file containing leads.\n"
            "Required columns: Name, Email\n"
            "Optional: Phone, Website",
            parse_mode='Markdown'
        )
        return UPLOAD_FILE
    elif query.data == "back_main":
        keyboard = [
            [InlineKeyboardButton("📄 Generate Invoice", callback_data="generate_invoice")],
            [InlineKeyboardButton("🔍 Lead Finder", callback_data="lead_finder")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(
            f"🏢 *Welcome to {COMPANY_NAME} Bot!*\n\n"
            "Choose an option:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return MAIN_MENU

async def niche_location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle niche and location input for lead finding"""
    user_input = update.message.text.strip()
    parts = user_input.split()
    if len(parts) < 2:
        await update.message.reply_text(
            "❌ Invalid format. Please enter in the format: *Niche Location*\n"
            "Example: Schools Lagos",
            parse_mode='Markdown'
        )
        return NICHE_LOCATION
    
    niche = parts[0]
    location = ' '.join(parts[1:])
    
    await update.message.reply_text(
        f"🔍 Searching for *{niche}* in *{location}*...\n\n"
        "This may take a few minutes.",
        parse_mode='Markdown'
    )
    
    # Run scraping in background or async
    leads = await find_business_leads(niche, location)
    
    if not leads:
        await update.message.reply_text(
            "❌ No leads found. Try a different search."
        )
        return ConversationHandler.END
    
    # Generate file
    file_path = generate_leads_file(leads)
    
    # Send file
    with open(file_path, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=f"leads_{niche}_{location}.xlsx",
            caption=f"✅ Found {len(leads)} leads for {niche} in {location}"
        )
    
    os.unlink(file_path)
    
    # Back to main menu
    keyboard = [
        [InlineKeyboardButton("🔄 New Search", callback_data="find_leads")],
        [InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_main")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "What would you like to do next?",
        reply_markup=reply_markup
    )
    return LEAD_FINDER_MENU

async def upload_file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle file upload for email sending"""
    if not update.message.document:
        await update.message.reply_text(
            "❌ Please upload a CSV or XLSX file."
        )
        return UPLOAD_FILE
    
    document = update.message.document
    file_name = document.file_name.lower()
    if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
        await update.message.reply_text(
            "❌ Please upload a CSV or XLSX file."
        )
        return UPLOAD_FILE
    
    # Download file
    file = await context.bot.get_file(document.file_id)
    file_path = f"temp_{update.effective_user.id}_{file_name}"
    await file.download_to_drive(file_path)
    
    # Store file path in context
    context.user_data['leads_file'] = file_path
    
    await update.message.reply_text(
        "✅ File uploaded successfully!\n\n"
        "Now, enter your *sender email address*:",
        parse_mode='Markdown'
    )
    return SENDER_EMAIL

async def sender_email_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle sender email input"""
    email = update.message.text.strip()
    # Basic email validation
    if '@' not in email or '.' not in email:
        await update.message.reply_text(
            "❌ Invalid email address. Please enter a valid email:"
        )
        return SENDER_EMAIL
    
    context.user_data['sender_email'] = email
    
    await update.message.reply_text(
        "✅ Sender email set!\n\n"
        "Enter your *email password* (or app password for Gmail):",
        parse_mode='Markdown'
    )
    return EMAIL_PASSWORD

async def email_password_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle email password input"""
    context.user_data['email_password'] = update.message.text.strip()
    
    await update.message.reply_text(
        "✅ Password set!\n\n"
        "Now, enter the *email subject*:",
        parse_mode='Markdown'
    )
    return EMAIL_SUBJECT

async def email_subject_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle email subject input"""
    context.user_data['email_subject'] = update.message.text.strip()
    
    await update.message.reply_text(
        "✅ Subject set!\n\n"
        "Now, enter the *email body* (message content):",
        parse_mode='Markdown'
    )
    return EMAIL_BODY

async def email_body_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle email body input and start sending"""
    context.user_data['email_body'] = update.message.text.strip()
    
    await update.message.reply_text(
        "✅ Email body set!\n\n"
        "Starting to send emails... (1 email every 30 minutes)",
        parse_mode='Markdown'
    )
    
    # Start sending emails in background
    asyncio.create_task(send_emails(update, context))
    
    return ConversationHandler.END

async def client_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle client name input"""
    user_id = update.effective_user.id
    if user_id not in user_data:
        user_data[user_id] = InvoiceData()
    
    user_data[user_id].client_name = update.message.text
    
    await update.message.reply_text(
        f"Client: *{update.message.text}*\n\n"
        "Now, tell me about the first item/service:\n"
        "Please enter the *service description*:",
        parse_mode='Markdown'
    )
    return SERVICE

async def service_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle service description"""
    context.user_data['current_service'] = update.message.text
    
    await update.message.reply_text(
        f"Service: *{update.message.text}*\n\n"
        "Enter the *price* (in Naira):",
        parse_mode='Markdown'
    )
    return PRICE

async def price_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle price input"""
    try:
        price = float(update.message.text)
        context.user_data['current_price'] = price
        
        await update.message.reply_text(
            f"Price: *{price:,.2f} Naira*\n\n"
            "Enter the *quantity*:",
            parse_mode='Markdown'
        )
        return QUANTITY
    except ValueError:
        await update.message.reply_text(
            "❌ Invalid price. Please enter a number (e.g., 5000):"
        )
        return PRICE

async def quantity_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle quantity input and add item"""
    try:
        quantity = int(update.message.text)
        user_id = update.effective_user.id
        
        # Add item to invoice
        service = context.user_data.get('current_service', '')
        price = context.user_data.get('current_price', 0)
        
        user_data[user_id].add_item(service, price, quantity)
        
        # Show item added confirmation
        item_total = price * quantity
        await update.message.reply_text(
            f"✅ *Item added successfully!*\n\n"
            f"Service: {service}\n"
            f"Quantity: {quantity}\n"
            f"Price: {price:,.2f} Naira\n"
            f"Item Total: {item_total:,.2f} Naira\n\n"
            f"Current total: {user_data[user_id].calculate_total():,.2f} Naira",
            parse_mode='Markdown'
        )
        
        # Ask if they want to add more
        keyboard = [
            [
                InlineKeyboardButton("✅ Yes, add another", callback_data="add_more"),
                InlineKeyboardButton("📄 No, generate invoice", callback_data="generate")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "Do you want to add another item?",
            reply_markup=reply_markup
        )
        return ADD_MORE
        
    except ValueError:
        await update.message.reply_text(
            "❌ Invalid quantity. Please enter a whole number (e.g., 2):"
        )
        return QUANTITY

async def add_more_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the add more / generate choice"""
    query = update.callback_query
    await query.answer()
    
    if query.data == "add_more":
        await query.edit_message_text(
            "Great! Enter the next *service description*:",
            parse_mode='Markdown'
        )
        return SERVICE
    else:
        # Ask for payment date
        await query.edit_message_text(
            "Now, let's add payment details.\n\n"
            "Enter the *payment date* (YYYY-MM-DD) [or 'today' for today's date]:",
            parse_mode='Markdown'
        )
        return PAYMENT_DATE

async def payment_date_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle payment date input"""
    user_id = update.effective_user.id
    date_text = update.message.text.lower()
    
    if date_text == 'today':
        payment_date = datetime.now().strftime("%Y-%m-%d")
    else:
        payment_date = date_text
        
    user_data[user_id].payment_date = payment_date
    
    await update.message.reply_text(
        f"Payment date: *{payment_date}*\n\n"
        "Enter *payment method* (Bank Transfer/Cash/USSD/etc):",
        parse_mode='Markdown'
    )
    return PAYMENT_METHOD

async def payment_method_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle payment method and generate invoice"""
    user_id = update.effective_user.id
    user_data[user_id].payment_method = update.message.text
    
    # Generate invoice number
    user_data[user_id].generate_invoice_number()
    
    await update.message.reply_text(
        "⏳ *Generating your invoice...*",
        parse_mode='Markdown'
    )
    
    # Generate PDF
    try:
        pdf_path = await generate_invoice_pdf(user_data[user_id])
        
        # Send PDF
        with open(pdf_path, 'rb') as pdf_file:
            await update.message.reply_document(
                document=pdf_file,
                filename=f"Tachaelhub_Invoice_{user_data[user_id].invoice_number}.pdf",
                caption=f"✅ *Invoice generated successfully!*\nInvoice #: {user_data[user_id].invoice_number}",
                parse_mode='Markdown'
            )
        
        # Clean up temp file
        os.unlink(pdf_path)
        
        # Ask to start over
        keyboard = [[InlineKeyboardButton("🔄 Create New Invoice", callback_data="new_invoice")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "What would you like to do next?",
            reply_markup=reply_markup
        )
        
        # Clear user data
        del user_data[user_id]
        
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        await update.message.reply_text(
            "❌ Sorry, there was an error generating the invoice. Please try again."
        )
    
    return ConversationHandler.END

async def new_invoice_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle new invoice button"""
    query = update.callback_query
    await query.answer()
    
    # Start new invoice
    user_id = update.effective_user.id
    user_data[user_id] = InvoiceData()
    
    await query.edit_message_text(
        "🏢 Starting new invoice...\n\nPlease enter the *client name*:",
        parse_mode='Markdown'
    )
    return CLIENT_NAME

async def generate_invoice_pdf(invoice_data):
    """Generate PDF invoice (unchanged)"""
    # Create temp file
    fd, path = tempfile.mkstemp(suffix='.pdf')
    os.close(fd)
    
    # Create PDF
    doc = SimpleDocTemplate(path, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=COLORS['deep_blue'],
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    # Company header
    header_text = f"""
    <font color='#1a237e' size='28'><b>{COMPANY_NAME}</b></font>
    """
    elements.append(Paragraph(header_text, title_style))
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Payment confirmation banner
    paid_text = """
    <para alignment='center'>
    <font color='#ffd700' size='16'><b>✓ PAYMENT RECEIPT / INVOICE ✓</b></font>
    </para>
    """
    elements.append(Paragraph(paid_text, styles['Normal']))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Invoice details
    details_data = [
        ['Invoice Number:', invoice_data.invoice_number, 'Payment Date:', invoice_data.payment_date],
        ['Client:', invoice_data.client_name, 'Payment Method:', invoice_data.payment_method],
        ['Status:', '✅ PAID', '', '']
    ]
    
    details_table = Table(details_data, colWidths=[80, 150, 80, 150])
    details_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), COLORS['deep_blue']),
        ('TEXTCOLOR', (2, 0), (2, -1), COLORS['deep_blue']),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(details_table)
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Items table header
    items_header = Paragraph(
        "<font color='#1a237e' size='12'><b>ITEMS/SERVICES PURCHASED</b></font>",
        styles['Normal']
    )
    elements.append(items_header)
    elements.append(Spacer(1, 0.1*inch))
    
    # Items table
    table_data = [['Item', 'Description', 'Qty', 'Unit Price (Naira)', 'Total (Naira)']]
    
    for i, item in enumerate(invoice_data.items, 1):
        total = item['price'] * item['quantity']
        table_data.append([
            str(i),
            item['service'],
            str(item['quantity']),
            f"{item['price']:,.2f}",
            f"{total:,.2f}"
        ])
    
    items_table = Table(table_data, colWidths=[40, 250, 40, 90, 90])
    items_table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), COLORS['deep_blue']),
        ('TEXTCOLOR', (0, 0), (-1, 0), COLORS['white']),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Table body style
        ('BACKGROUND', (0, 1), (-1, -1), COLORS['white']),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        
        # Grid lines
        ('GRID', (0, 0), (-1, -1), 1, COLORS['deep_blue']),
        ('LINEBELOW', (0, 0), (-1, 0), 2, COLORS['gold']),
    ]))
    elements.append(items_table)
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Total
    total_amount = invoice_data.calculate_total()
    
    # Totals table
    totals_data = [
        ['Subtotal:', f"{total_amount:,.2f} Naira"],
        ['', ''],
        ['TOTAL AMOUNT PAID:', f"{total_amount:,.2f} Naira"],
    ]
    
    totals_table = Table(totals_data, colWidths=[400, 150])
    totals_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTSIZE', (0, 2), (1, 2), 14),
        ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 2), (1, 2), COLORS['deep_blue']),
        ('BACKGROUND', (0, 2), (1, 2), COLORS['gold']),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(totals_table)
    
    elements.append(Spacer(1, 0.5*inch))
    
    # Footer
    footer_text = f"""
    <para alignment='center'>
    <font color='#1a237e' size='10'>
    <b>{COMPANY_NAME}</b> - Official Payment Receipt<br/>
    This document serves as proof of payment. Thank you for your business!<br/>
    For inquiries, please reference invoice: {invoice_data.invoice_number}
    </font>
    </para>
    """
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    return path

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancel conversation"""
    user_id = update.effective_user.id
    if user_id in user_data:
        del user_data[user_id]
    
    await update.message.reply_text(
        "❌ Invoice generation cancelled. Use /start to begin again."
    )
    return ConversationHandler.END

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Help command"""
    help_text = (
        "🤖 *Tachaelhub Bot Commands:*\n\n"
        "/start - Start and choose between Invoice Generator or Lead Finder\n"
        "/help - Show this help message\n"
        "/cancel - Cancel current operation\n\n"
        "*Invoice Generator:*\n"
        "1. Start with /start\n"
        "2. Choose Generate Invoice\n"
        "3. Enter client name\n"
        "4. Add item(s) with description, price, quantity\n"
        "5. Add payment date and method\n"
        "6. Receive PDF invoice\n\n"
        "*Lead Finder:*\n"
        "1. Choose Lead Finder from main menu\n"
        "2. Find Business Leads: Enter niche and location\n"
        "3. Send Outreach Emails: Upload CSV/XLSX file\n"
        "4. Enter email details and send"
    )
    await update.message.reply_text(help_text, parse_mode='Markdown')

# Lead Finder Functions
async def find_business_leads(niche, location):
    """Find business leads by scraping Google and Google Maps"""
    leads = []
    
    # Search Google for businesses
    query = f"{niche} in {location}"
    try:
        search_results = list(search(query, num=20, stop=20, pause=2))
    except Exception as e:
        logger.error(f"Search error: {e}")
        return leads
    
    for url in search_results:
        lead = None
        if 'maps.google.com' in url or 'google.com/maps' in url:
            # Extract from Maps URL
            lead = await extract_from_maps_url(url)
        else:
            # Try to extract business info from regular search result
            lead = await extract_from_search_url(url, niche, location)
        
        if lead:
            leads.append(lead)
            if len(leads) >= 10:  # Limit to 10 leads
                break
    
    # Classify leads
    for lead in leads:
        if lead.get('website') and lead['website'] != url:
            # Scrape website for email
            email = await scrape_email_from_website(lead['website'])
            if email:
                lead['email'] = email
            lead['status'] = 'HAS WEBSITE'
        elif 'facebook.com' in str(lead.get('website', '')):
            lead['status'] = 'FACEBOOK ONLY'
        else:
            lead['status'] = 'NO WEBSITE'
    
    return leads

async def extract_from_maps_url(url):
    """Extract business info from Google Maps URL"""
    # Parse URL to get business name
    # URL format: https://www.google.com/maps/place/Business+Name/@lat,lng,zoomz/data=!3m1!4b1!4m6!3m5!1s...!8m2!3dlat!4dlng!16s...
    try:
        # Extract name from URL
        if '/place/' in url:
            name_part = url.split('/place/')[1].split('/')[0].replace('+', ' ')
            name = name_part.split('@')[0]
        else:
            name = "Business from Maps"
        
        return {
            'name': name,
            'phone': None,  # Would need API or scraping
            'email': None,
            'website': None,
            'maps_link': url,
            'status': 'NO WEBSITE'  # Will be updated later
        }
    except:
        return None

async def extract_from_search_url(url, niche, location):
    """Extract from regular search URL"""
    try:
        response = requests.get(url, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extract name
        title = soup.find('title')
        name = title.text.strip() if title else f"{niche} in {location}"
        
        # Extract phone
        phone_pattern = r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b|\b\d{4}[-.]?\d{3}[-.]?\d{4}\b'
        phone_match = re.search(phone_pattern, soup.text)
        phone = phone_match.group() if phone_match else None
        
        # Extract email
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        email_match = re.search(email_pattern, soup.text)
        email = email_match.group() if email_match else None
        
        # Website is the URL
        website = url
        
        return {
            'name': name,
            'phone': phone,
            'email': email,
            'website': website,
            'maps_link': f"https://maps.google.com/?q={name.replace(' ', '+')}+{location.replace(' ', '+')}",
            'status': 'HAS WEBSITE' if website else 'NO WEBSITE'
        }
    except Exception as e:
        logger.error(f"Error extracting from {url}: {e}")
        return None

async def scrape_email_from_website(url):
    """Scrape email from website"""
    try:
        response = requests.get(url, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Look for mailto links
        mailto_links = soup.find_all('a', href=re.compile(r'mailto:'))
        if mailto_links:
            email = mailto_links[0]['href'].replace('mailto:', '')
            return email
        
        # Look for email in text
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        match = re.search(email_pattern, soup.text)
        if match:
            return match.group()
        
        # Try contact page
        contact_links = soup.find_all('a', href=re.compile(r'contact', re.I))
        for link in contact_links[:3]:  # Check first 3
            contact_url = link['href']
            if not contact_url.startswith('http'):
                contact_url = url.rstrip('/') + '/' + contact_url.lstrip('/')
            try:
                contact_response = requests.get(contact_url, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
                contact_soup = BeautifulSoup(contact_response.text, 'html.parser')
                match = re.search(email_pattern, contact_soup.text)
                if match:
                    return match.group()
            except:
                continue
        
        return None
    except Exception as e:
        logger.error(f"Error scraping {url}: {e}")
        return None
        phone_pattern = r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b'
        phone_match = re.search(phone_pattern, soup.text)
        phone = phone_match.group() if phone_match else None
        
        # Extract email
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        email_match = re.search(email_pattern, soup.text)
        email = email_match.group() if email_match else None
        
        # Extract website
        website = url
        
        return {
            'name': name,
            'phone': phone,
            'email': email,
            'website': website,
            'maps_link': f"https://maps.google.com/?q={name} {location}",
            'status': 'HAS WEBSITE' if website else 'NO WEBSITE'
        }
    except Exception as e:
        logger.error(f"Error extracting from {url}: {e}")
        return None

def generate_leads_file(leads):
    """Generate XLSX file from leads"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Headers
    headers = ['Name', 'Phone', 'Email', 'Website', 'Maps_Link', 'Status']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Data
    for row_num, lead in enumerate(leads, 2):
        ws.cell(row=row_num, column=1, value=lead.get('name', ''))
        ws.cell(row=row_num, column=2, value=lead.get('phone', ''))
        ws.cell(row=row_num, column=3, value=lead.get('email', ''))
        ws.cell(row=row_num, column=4, value=lead.get('website', ''))
        ws.cell(row=row_num, column=5, value=lead.get('maps_link', ''))
        ws.cell(row=row_num, column=6, value=lead.get('status', ''))
    
    file_path = tempfile.mktemp(suffix='.xlsx')
    wb.save(file_path)
    return file_path

async def send_emails(update, context):
    """Send emails from file with delays"""
    file_path = context.user_data.get('leads_file')
    sender_email = context.user_data.get('sender_email')
    password = context.user_data.get('email_password')
    subject = context.user_data.get('email_subject')
    body = context.user_data.get('email_body')
    
    if not file_path or not os.path.exists(file_path):
        await update.message.reply_text("❌ File not found.")
        return
    
    # Read file
    if file_path.endswith('.csv'):
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            df = list(reader)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        df = []
        for row in ws.iter_rows(min_row=2):
            row_dict = {}
            for col_num, cell in enumerate(row):
                row_dict[headers[col_num]] = cell.value
            df.append(row_dict)
    
    # Check required columns
    if 'Email' not in df[0] or 'Name' not in df[0]:
        await update.message.reply_text("❌ File must have 'Name' and 'Email' columns.")
        return
    
    total = len(df)
    sent = 0
    
    # SMTP setup
    smtp_server = 'smtp.gmail.com' if 'gmail.com' in sender_email else 'smtp.mail.yahoo.com'  # Add more if needed
    smtp_port = 587
    
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    try:
        server.login(sender_email, password)
    except Exception as e:
        await update.message.reply_text(f"❌ SMTP login failed: {e}")
        return
    
    for row in df:
        email = row.get('Email')
        name = row.get('Name')
        
        if not email or str(email).lower() in ('nan', 'none', ''):
            continue
        
        # Personalize body
        personalized_body = body.replace('{name}', str(name) if name else '')
        
        try:
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = email
            msg['Subject'] = subject
            msg.attach(MIMEText(personalized_body, 'plain'))
            
            server.sendmail(sender_email, email, msg.as_string())
            sent += 1
            
            # Update progress
            await update.message.reply_text(f"Sending email {sent} of {total}")
            
            # Wait 30 minutes (1800 seconds) except for last
            if sent < total:
                await asyncio.sleep(1800)
        
        except Exception as e:
            logger.error(f"Error sending to {email}: {e}")
    
    server.quit()
    await update.message.reply_text(f"✅ Sent {sent} emails out of {total}.")
    
    # Clean up
    os.unlink(file_path)

async def main():
    """Main function to run the bot"""
    global application
    
    # Get token from environment variable
    TOKEN = os.environ.get('BOT_TOKEN', "YOUR_BOT_TOKEN_HERE")
    
    # Start HTTP server for health checks and webhooks
    await start_http_server()
    
    # Create application
    application = Application.builder().token(TOKEN).build()
    
    # Create conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            MAIN_MENU: [CallbackQueryHandler(main_menu_handler, pattern="^(generate_invoice|lead_finder)$")],
            LEAD_FINDER_MENU: [CallbackQueryHandler(lead_finder_menu_handler, pattern="^(find_leads|send_emails|back_main)$")],
            CLIENT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, client_name_handler)],
            SERVICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, service_handler)],
            PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, price_handler)],
            QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, quantity_handler)],
            ADD_MORE: [CallbackQueryHandler(add_more_handler, pattern="^(add_more|generate)$")],
            PAYMENT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, payment_date_handler)],
            PAYMENT_METHOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, payment_method_handler)],
            NICHE_LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, niche_location_handler)],
            UPLOAD_FILE: [MessageHandler(filters.Document.ALL, upload_file_handler)],
            SENDER_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, sender_email_handler)],
            EMAIL_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, email_password_handler)],
            EMAIL_SUBJECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, email_subject_handler)],
            EMAIL_BODY: [MessageHandler(filters.TEXT & ~filters.COMMAND, email_body_handler)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    application.add_handler(conv_handler)
    application.add_handler(CommandHandler('help', help_command))
    application.add_handler(CallbackQueryHandler(new_invoice_callback, pattern="^new_invoice$"))
    
    # Initialize and start the application
    await application.initialize()
    await application.start()
    
    # Check for webhook URL
    WEBHOOK_URL = os.environ.get('WEBHOOK_URL')
    if WEBHOOK_URL:
        await application.bot.set_webhook(url=WEBHOOK_URL)
        logger.info(f"Webhook set to {WEBHOOK_URL}")
        
        # Keep the application running
        try:
            while True:
                await asyncio.sleep(3600)  # Sleep for an hour
        except KeyboardInterrupt:
            await application.stop()
    else:
        # Fallback to polling for local development
        logger.info("Starting in polling mode")
        await application.run_polling()

if __name__ == '__main__':
    asyncio.run(main())